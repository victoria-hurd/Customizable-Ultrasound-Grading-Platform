import sys
import os
import pandas as pd
from PyQt6.QtWidgets import (
    QApplication, QWidget, QFileDialog, QScrollArea, QVBoxLayout, 
    QLabel, QRadioButton, QButtonGroup, QFileDialog
    )
from PyQt6.QtMultimedia import QMediaPlayer, QVideoFrameFormat
from PyQt6.QtMultimediaWidgets import QVideoWidget
from PyQt6.QtCore import QUrl
from PyQt6 import uic
from openpyxl import Workbook

def getParticipantData():
        participantID = 1
        missionID = "EA-100"
        EVA_ID = 4
        return participantID, missionID, EVA_ID

def getGraderData():
        graderID = "VH"
        condition = "Post-Grade"
        return graderID, condition

class GradingApp(QWidget):
    def __init__(self):
        super().__init__()
        #uic.loadUi('gui.ui', self)
        self.setWindowTitle("Demo Ultrasound Grading Application")

        self.mediaPlayer = QMediaPlayer()       
        self.videoWidget = QVideoWidget()
        self.mediaPlayer.setVideoOutput(self.videoWidget)

        #self.playButton.clicked.connect(self.playVideo)
        #self.pauseButton.clicked.connect(self.pauseVideo)
        #self.openButton.clicked.connect(self.loadNewExam)
        #self.saveDataButton.clicked.connect(self.saveData)

        #self.openButton.clicked.connect(self.select_excel_and_load)

        main_layout = QVBoxLayout()


        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        self.button_groups = []

        excel_path = self.get_excel_path()
        questions = self.load_questions(excel_path)

        for q_text, options in questions:
            scroll_layout.addWidget(QLabel(q_text))

            group = QButtonGroup(self)
            self.button_groups.append(group)

            for opt in options:
                rb = QRadioButton(opt)
                group.addButton(rb)
                scroll_layout.addWidget(rb)

        scroll_widget.setLayout(scroll_layout)
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)

        main_layout.addWidget(scroll)
        self.setLayout(main_layout)

    def clear_layout(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

    def load_survey_questions(self, excel_path):
        df = pd.read_excel(excel_path)

        self.button_groups = []

        for _, row in df.iterrows():
            question_text = row["Question"]
            options = [opt for opt in row[1:] if pd.notna(opt)]

            # Question label
            label = QLabel(question_text)
            label.setStyleSheet("font-weight: bold; margin-top: 12px;")
            self.verticalLayout.addWidget(label)

            # Radio buttons
            group = QButtonGroup(self)
            self.button_groups.append(group)

            for opt in options:
                rb = QRadioButton(str(opt))
                group.addButton(rb)
                self.verticalLayout.addWidget(rb)

    def select_excel_and_load(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Survey File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        self.clear_layout(self.verticalLayout)
        self.load_survey_questions(path)

    def load_questions(self, excel_path):
        df = pd.read_excel(excel_path)
        questions = []

        for _, row in df.iterrows():
            question = row["Question"]
            options = [opt for opt in row[1:] if pd.notna(opt)]
            questions.append((question, options))

        return questions

    def get_excel_path(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Grading Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        return file_path

    def loadNewExam(self):
        self.successfulSaveLabel.setEnabled(False)
        filePath, _ = QFileDialog.getOpenFileName(self, "Choose Video File",
                ".", "Video Files (*.mp4 *.flv *.ts *.mts *.avi)")
        
        if os.path.exists(filePath):
            print(f"The file '{filePath}' exists.")
        else:
            print(f"The file '{filePath}' does not exist.")

        if filePath != '':
            self.mediaPlayer.setSource(QUrl.fromLocalFile(filePath))
            self.videoWidget.resize(640, 480)
            self.loadedFileLabel.setText(filePath)

    def playVideo(self):
        print("Trying to play video...")
        self.videoWidget.show()
        self.mediaPlayer.play()

    def pauseVideo(self):
        print("Trying to pause video...")
        self.mediaPlayer.pause()

    def checkAllButtonGroupsPushed():
        # Function to call in the beginning of saveData to ensure all button groups have a selection

        pass
        
    def saveData(self):
        print("Saving data...")
        participantID, missionID, EVA_ID = getParticipantData()
        graderID, condition = getGraderData()
        # Create a new workbook and select the active sheet
        workbook = Workbook()
        sheet = workbook.active
        # If sheet does not yet exist, create it and save headers
        headers = ["ParticipantID", "MissionID", "EVA_ID", "GraderID", "Condition",
                   "Rest1", "RightFlex", "LeftFlex", "Rest2"]
        sheet.append(headers)
        # Define your variables as a list
        variables = [participantID, missionID, EVA_ID, graderID, condition,
                     self.buttonGroupRest1.checkedButton().text(),
                     self.buttonGroupRightFlex.checkedButton().text(),
                     self.buttonGroupLeftFlex.checkedButton().text(),
                     self.buttonGroupRest2.checkedButton().text()]
        # Write the variables to the first row
        sheet.append(variables)
        # Save the workbook
        directory = os.getcwd()
        appDataDir = "App Data"
        outName = f"{graderID}_{missionID}_EVA{EVA_ID}_Grades.xlsx"
        outPath = os.path.join(directory,appDataDir,outName)
        workbook.save(outPath)
        # To load existing workbook, 
        # from openpyxl import load_workbook
        # wb = load_workbook("your_excel_file.xlsx")
        # wb.save("new_excel_file.xlsx")

        self.successfulSaveLabel.setEnabled(True)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    myapp = GradingApp()
    myapp.show()

    try:
        sys.exit(app.exec())
    except SystemExit:
        print("Closing Ultrasound Grading Application...")


# if __name__ == '__main__':
#     app = QApplication(sys.argv)
#     main_win = MainWindow()
#     available_geometry = main_win.screen().availableGeometry()
#     main_win.resize(available_geometry.width() / 3,
#                     available_geometry.height() / 2)
#     main_win.show()
#     sys.exit(app.exec())
