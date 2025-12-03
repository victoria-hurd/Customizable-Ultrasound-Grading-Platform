import sys
import os
from PyQt6.QtWidgets import QApplication, QWidget, QFileDialog
from PyQt6.QtMultimedia import QMediaPlayer, QVideoFrameFormat
from PyQt6.QtMultimediaWidgets import QVideoWidget
from PyQt6.QtCore import QUrl
from PyQt6 import uic
from openpyxl import Workbook

def getParticipantData():
        participantID = 1
        missionID = "EA-100"
        EVA_ID = 4
        return participantID, missionID, EVA_ID

def getGraderData():
        graderID = "VH"
        condition = "Post-Grade"
        return graderID, condition

class MyApp(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi('gui.ui', self)
        self.setWindowTitle("Demo Ultrasound Grading Application")

        self.mediaPlayer = QMediaPlayer()       
        self.videoWidget = QVideoWidget()
        self.mediaPlayer.setVideoOutput(self.videoWidget)

        self.playButton.clicked.connect(self.playVideo)
        self.pauseButton.clicked.connect(self.pauseVideo)
        self.openButton.clicked.connect(self.loadNewExam)
        self.saveDataButton.clicked.connect(self.saveData)


    def loadNewExam(self):
        self.successfulSaveLabel.setEnabled(False)
        filePath, _ = QFileDialog.getOpenFileName(self, "Choose Video File",
                ".", "Video Files (*.mp4 *.flv *.ts *.mts *.avi)")
        
        if os.path.exists(filePath):
            print(f"The file '{filePath}' exists.")
        else:
            print(f"The file '{filePath}' does not exist.")

        if filePath != '':
            self.mediaPlayer.setSource(QUrl.fromLocalFile(filePath))
            self.videoWidget.resize(640, 480)
            self.loadedFileLabel.setText(filePath)

    def playVideo(self):
        print("Trying to play video...")
        self.videoWidget.show()
        self.mediaPlayer.play()

    def pauseVideo(self):
        print("Trying to pause video...")
        self.mediaPlayer.pause()

    def checkAllButtonGroupsPushed():
        # Function to call in the beginning of saveData to ensure all button groups have a selection

        pass
        

    def saveData(self):
        print("Saving data...")
        participantID, missionID, EVA_ID = getParticipantData()
        graderID, condition = getGraderData()
        # Create a new workbook and select the active sheet
        workbook = Workbook()
        sheet = workbook.active
        # If sheet does not yet exist, create it and save headers
        headers = ["ParticipantID", "MissionID", "EVA_ID", "GraderID", "Condition",
                   "Rest1", "RightFlex", "LeftFlex", "Rest2"]
        sheet.append(headers)
        # Define your variables as a list
        variables = [participantID, missionID, EVA_ID, graderID, condition,
                     self.buttonGroupRest1.checkedButton().text(),
                     self.buttonGroupRightFlex.checkedButton().text(),
                     self.buttonGroupLeftFlex.checkedButton().text(),
                     self.buttonGroupRest2.checkedButton().text()]
        # Write the variables to the first row
        sheet.append(variables)
        # Save the workbook
        directory = os.getcwd()
        appDataDir = "App Data"
        outName = f"{graderID}_{missionID}_EVA{EVA_ID}_Grades.xlsx"
        outPath = os.path.join(directory,appDataDir,outName)
        workbook.save(outPath)
        # To load existing workbook, 
        # from openpyxl import load_workbook
        # wb = load_workbook("your_excel_file.xlsx")
        # wb.save("new_excel_file.xlsx")

        self.successfulSaveLabel.setEnabled(True)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    myapp = MyApp()
    myapp.show()

    try:
        sys.exit(app.exec())
    except SystemExit:
        print("Closing Ultrasound Grading Application...")


# if __name__ == '__main__':
#     app = QApplication(sys.argv)
#     main_win = MainWindow()
#     available_geometry = main_win.screen().availableGeometry()
#     main_win.resize(available_geometry.width() / 3,
#                     available_geometry.height() / 2)
#     main_win.show()
#     sys.exit(app.exec())
